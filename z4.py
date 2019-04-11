# -*- coding:utf-8 -*-
# d={"a":1,"b":2}
# for i,j in (d.items()):
#     print(i,j)
# for i1 in d.keys():
#     print(i1)
# from optparse import OptionParser
# class A:
#     def __init__(self):
#         self.op= ()
#         self.op.add_option("-s",dest="server")
#         option,args=self.op.parse_args()
#         print(option)
#         print(args)
#         self.vertify(option,args)
#     def vertify(self,option,args):
#         print(option,args)
#         if hasattr(self,args[0]):
#             getattr(self,args[0])()
#     def run(self):
#         print(88)
# a=A()



# import requests
# ret=requests.get("http://39.105.25.57/csk.csv")
# data=ret.content
# print(type(data))
# # print(data)
# with open("csk.csv","w",e) as code:
#     code.write(data.decode())


# with open("cs1.csv","w",encoding="utf-8") as w:
#     w.write("姓名,年龄,性别")
# from openpyxl import Workbook
# wb=Workbook()
# ws=wb.active
# m = [[1, 2, 3],  [4, 5, 6],  [7, 8, 9]]
# n = [["a", "b", "c"],  ["d", "e", "f"],  ["g", "h", "i"]]
# # 矩阵点乘
# zip_list=list(zip(m,n))
# for i,j in zip_list:
#     ws.append(i)
#     ws.append(j)
# # ws.append((list(zip(m,n))
# # ws.append([1,2,3])
# wb.save("array.xlsx")

# def money_format(value):
#     value = "%.2f" % float(value)
#     components = str(value).split('.')
#     if len(components) > 1:
#         left, right = components
#         right = '.' + right
#     else:
#         left, right = components[0], ''
#
#     result = ''
#     print(left)
#     while left:
#         result = left[-3:] + ',' + result
#         left = left[:-3]
#     return result.strip(',') + right
#
# ret=money_format(15000.2)
# print(ret)
# print(100*"*")
# print(round(12.36,2))
# print(float("%0.2f" % 1.2))
# s="ab,c"
# print(",".join(["1","2","3"]))
# print(s.split(","))
# students = [
#         ('john', 'A', 15,1),
#         ('johnn', 'A', 15,5),
#         ('johnn', 'A', 15,4),
#         ('jane', 'B', 12,2),
#         ('dave', 'B', 10,3),
# ]
# from operator import itemgetter, attrgetter
# print(itemgetter(2,1)(students[2]))




from collections import Iterable, Iterator


# class Stu:
#     def __init__(self):
#         self.item = [1,3,5]
#         self.nane = "li"
#     def __iter__(self):
#         iter = Iter(self.nane)
#         return iter
#
# class Iter:
#     def __init__(self,*args):
#         self.diedai = args[0]
#         self.current_index = 0
#     def __iter__(self):
#         return self
#     def next(self):
#         if self.current_index < len(self.diedai):
#             self.current_index += 1
#             return self.diedai[self.current_index-1]
#         else:
#             print("迭代完成")
#             raise StopIteration
#
# stu = Stu()
# print(iter(stu))
# iter函数 调用可迭代对象的__iter__方法 接收到迭代器

# for i in stu:
#     print(i)
# 遍历可迭代对象  本质是先取迭代器 在用迭代器取一次取值
# 需要在__next__抛出停止迭代异常遍历完成之后才会停止

# sit = Iter(stu.item)
# print(sit)
# 实例化迭代器  实例时需要接受迭代对象的可遍历属性  而不是迭代对象
# print(next(sit))
# print(next(sit))
# next函数调用迭代器对象的__next__方法 返回可迭代对象属性的第一条数据及下一条数据
# while True:
#     try:
#         print(next(sit))
#     except StopIteration as e:
#         break
# 结束时需捕获异常

# def fib():
#     n,a,b = 0,0,1
#     while n < 10:
#         # print(b)
#         yield b
#         print(666)
#         a, b = b, a + b
#         n+=1
#
# print(next(fib()))
# print(next(fib()))
# print(next(fib()))
import sys
reload(sys)
print(sys.set())
from openpyxl import load_workbook,Workbook

# wb=Workbook()
# ws=wb.active
# ws.title="员工"
# ws.append(["姓名","性别","年龄"])
# ws.append(["小李","男","18"])
# ws.append(["小花","女","18"])


# ws["H4"]="帅"
# ws.merge_cells("H4:I7")
#
# ws1=wb.create_sheet("申请单",1)
# ws1.append(["申请人","审核人"])
# wb.save("sheet.xlsx")

# from openpyxl import load_workbook
wb = load_workbook("sheet.xlsx")
print(wb.sheetnames)
sheet = wb.get_sheet_by_name("员工")
for i in [sheet["A1"],sheet["B1"]]:
    print(i.value)
# print(sheet["B2"].value)


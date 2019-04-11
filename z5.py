import pandas as pd
from openpyxl import load_workbook
data = [{'name':'ls','age':10},{'gender':'men','age':20},{'height':180}]
dt = pd.DataFrame(data)
dt=dt[['name','gender']]
dt.columns = ['姓名','性别']
excel_path= '{}.xlsx'.format('oa')
excelWriter = pd.ExcelWriter(excel_path,engine='openpyxl')
dt.to_excel(excel_writer=excelWriter,sheet_name='表一',encoding='utf-8',index=None)
excelWriter.save()

dt1=pd.DataFrame(data)
dt1=dt1[['height']]
dt1.columns=['身高']
# wb = load_workbook(excel_path)
wb = load_workbook(excelWriter.path)
excelWriter.book=wb
dt1.to_excel(excel_writer=excelWriter,sheet_name='表二',index=None)
excelWriter.close()







